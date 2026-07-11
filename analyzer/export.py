import re
import openpyxl
from openpyxl.utils import get_column_letter
from analyzer.database import db_session
from analyzer.constants import MatchStatus, CategoryType
from analyzer.logging_config import logger
from analyzer import repository


_INVALID_SHEET_CHARS = re.compile(r'[:\\/?*\[\]]')


def _safe_sheet_title(title, used_titles):
    """Excel sheet names: <=31 chars, no : \\ / ? * [ ], and unique within
    the workbook. Account names are free text (see FileImportSettingsDialog)
    so this sanitizes/dedupes rather than trusting them blindly."""
    cleaned = _INVALID_SHEET_CHARS.sub("-", title)[:31]
    candidate = cleaned
    n = 2
    while candidate in used_titles:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used_titles.add(candidate)
    return candidate


# Fields useful day-to-day vs. fields only useful for troubleshooting a
# specific row (which rule/match/import produced it, its raw txn_id).
# System fields are still exported (nothing is thrown away) but hidden
# by default so they don't clutter the sheet.
CONSOLIDATED_SYSTEM_FIELDS = {"category_src", "rule_id", "match_id", "import_id", "txn_id"}
ACCOUNT_SYSTEM_FIELDS = {"category_src", "rule_id", "match_id", "import_id", "txn_id", "source_file", "source_row"}
TRANSFER_SYSTEM_FIELDS = {"match_id"}


def get_export_summary():
    with db_session(commit=False) as conn:
        total_transactions = conn.execute("SELECT COUNT(*) AS count FROM transactions").fetchone()["count"]
        uncategorized = conn.execute(
            "SELECT COUNT(*) AS count FROM transactions WHERE category IS NULL OR category=''"
        ).fetchone()["count"]
        accepted_transfers = conn.execute(
            "SELECT COUNT(*) AS count FROM matches m JOIN transactions d ON m.debit_txn = d.txn_id JOIN transactions c ON m.credit_txn = c.txn_id WHERE m.status = ?",
            (MatchStatus.ACCEPTED.value,),
        ).fetchone()["count"]
        return {
            "total_transactions": total_transactions,
            "uncategorized": uncategorized,
            "accepted_transfers": accepted_transfers,
        }


def export_workbook(output_path: str):
    with db_session(commit=False) as conn:
        total_transactions = conn.execute("SELECT COUNT(*) AS count FROM transactions").fetchone()["count"]
        if total_transactions <= 0:
            raise ValueError("Nothing to export. Import at least one transaction first.")

    with db_session(commit=False) as conn:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        used_titles = {"Consolidated", "Self Transfers", "Uncategorized",
                       "Category Summary", "Category Type Summary"}

        def add_sheet(title, query, params=None, hidden_fields=None):
            ws = wb.create_sheet(title=title)
            rows = conn.execute(query, params or []).fetchall()
            if not rows:
                logger.warning(f"No data for sheet {title}. Skipping.")
                wb.remove(ws)  # previously left an empty sheet behind
                return None
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row[h] for h in headers])

            if hidden_fields:
                for col_idx, header in enumerate(headers, 1):
                    if header in hidden_fields:
                        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
            return ws

        # Consolidated: one row per transaction across all banks/accounts.
        # category_type now joined in so income/expense/transfer reporting
        # doesn't require opening a separate sheet.
        add_sheet("Consolidated",
                """SELECT t.txn_date, t.bank, t.account, t.description, t.amount, t.dr_cr,
                          t.category, ct.category_type,
                          t.category_src, t.rule_id, t.match_id, t.import_id, t.txn_id
                   FROM transactions t
                   LEFT JOIN category_types ct ON t.category = ct.category
                   ORDER BY t.txn_date DESC""",
                hidden_fields=CONSOLIDATED_SYSTEM_FIELDS)

        # Per-account instead of per-bank: two accounts at the same bank
        # (e.g. "Primary Savings" and "Business Account", both HDFC) now
        # get their own sheet instead of being lumped into one "Bank -
        # HDFC" sheet.
        accounts = [row[0] for row in conn.execute("SELECT DISTINCT account FROM transactions ORDER BY account").fetchall()]
        for account in accounts:
            add_sheet(_safe_sheet_title(f"Account - {account}", used_titles),
                    """SELECT t.txn_date, t.bank, t.account, t.description,
                              CASE WHEN t.dr_cr='DR' THEN t.amount END AS debit,
                              CASE WHEN t.dr_cr='CR' THEN t.amount END AS credit,
                              t.balance, t.reference, t.payment_mode,
                              t.category, ct.category_type,
                              t.category_src, t.rule_id, t.match_id, t.import_id, t.txn_id,
                              t.source_file, t.source_row
                       FROM transactions t
                       LEFT JOIN category_types ct ON t.category = ct.category
                       WHERE t.account=? ORDER BY t.txn_date DESC""",
                    (account,),
                    hidden_fields=ACCOUNT_SYSTEM_FIELDS)

        add_sheet("Self Transfers",
                """SELECT d.txn_date, d.amount, d.account AS from_account, c.account AS to_account,
                          d.description AS debit_desc, c.description AS credit_desc,
                          m.reason, m.match_id
                   FROM matches m
                   JOIN transactions d ON m.debit_txn = d.txn_id
                   JOIN transactions c ON m.credit_txn = c.txn_id
                   WHERE m.status = ?
                   ORDER BY d.txn_date DESC""",
                (MatchStatus.ACCEPTED.value,),
                hidden_fields=TRANSFER_SYSTEM_FIELDS)

        add_sheet("Uncategorized",
                """SELECT txn_date, bank, account, description, amount, dr_cr, txn_id
                   FROM transactions
                   WHERE category IS NULL OR category=''
                   ORDER BY txn_date DESC""",
                hidden_fields={"txn_id"})

        add_sheet("Category Summary",
                "SELECT category, dr_cr, COUNT(*) as count, SUM(amount) as total " \
                "   FROM transactions " \
                "   WHERE category IS NOT NULL GROUP BY category, dr_cr")

        add_sheet("Category Type Summary",
                """SELECT ct.category_type, t.dr_cr, COUNT(*) as count, SUM(t.amount) as total
                    FROM transactions t
                    JOIN category_types ct ON t.category = ct.category
                    WHERE t.category IS NOT NULL AND t.category != ''
                    GROUP BY ct.category_type, t.dr_cr
                    ORDER BY ct.category_type, t.dr_cr""")

        wb.save(output_path)
        logger.info(f"Workbook saved to {output_path}")