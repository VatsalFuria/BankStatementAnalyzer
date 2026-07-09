import openpyxl
from analyzer.parsers.base_parser import BaseParser
from analyzer.models import StandardTransaction
from analyzer.utils import parse_amount
from analyzer.exceptions import ParseError
from analyzer.logging_config import logger

class ExampleComplexBankParser(BaseParser):
    """
    Template for statements that don't fit ConfigurableExcelParser's
    assumption of one clean header row: merged title cells, a logo/address
    block above the real header, or columns that shift position between
    exports. Copy this file, rename the class, and adjust HEADER_MARKERS
    and the column lookups for your bank.
    """
    display_name = "Example Complex Bank (template)"
    HEADER_MARKERS = ("Date", "Narration")

    def can_parse(self, filepath: str) -> bool:
        if not filepath.lower().endswith((".xlsx", ".xls")):
            return False
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            return self._find_header_row(wb.active) is not None
        except Exception:
            logger.exception(f"can_parse probing failed for {filepath}")
            return False

    def _find_header_row(self, ws):
        # Scan the first 15 rows instead of assuming row 1 — handles
        # logo/title rows sitting above the real header.
        for row in ws.iter_rows(min_row=1, max_row=15):
            values = [str(c.value).strip() if c.value else "" for c in row]
            if all(m in values for m in self.HEADER_MARKERS):
                return row[0].row
        return None

    def parse(self, filepath: str):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Unmerge and forward-fill merged ranges so the rest of the logic
        # can treat the sheet as a plain grid.
        for merged in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged.bounds
            top_left = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left

        header_row_num = self._find_header_row(ws)
        if header_row_num is None:
            raise ParseError("Could not locate header row", filepath=filepath)
        headers = {cell.value: idx for idx, cell in enumerate(ws[header_row_num])}

        transactions, skipped = [], []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_num + 1), start=header_row_num + 1):
            values = [c.value for c in row]
            if not values or all(v in (None, "") for v in values):
                continue  # blank spacer row
            try:
                withdraw = parse_amount(values[headers["Withdrawal Amt."]], filepath=filepath, row=row_idx, column="Withdrawal Amt.")
                deposit = parse_amount(values[headers["Deposit Amt."]], filepath=filepath, row=row_idx, column="Deposit Amt.")
                transactions.append(StandardTransaction(
                    bank="ExampleBank", account="",
                    txn_date=str(values[headers["Date"]]),
                    description=str(values[headers["Narration"]] or ""),
                    amount=withdraw if withdraw > 0 else deposit,
                    dr_cr="DR" if withdraw > 0 else "CR",
                    source_file=filepath, source_row=row_idx,
                ))
            except ParseError as e:
                logger.warning(str(e))
                skipped.append((row_idx, str(e)))

        self.last_skipped_rows = skipped
        return transactions